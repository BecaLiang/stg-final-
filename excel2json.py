import os
import uuid
import json
import shutil
import argparse
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import pandas as pd
import io
from PIL import Image as PILImage

def generate_uuid():
    """Generate a UUID for entity IDs"""
    return str(uuid.uuid4())

def identify_template_type(workbook):
    """
    Determine which template type the Excel file matches
    Returns: "CEA", "EQ", "starteam", or None if no match
    """
    # Check sheet names
    sheet_names = workbook.sheetnames
    
    # Get first sheet for analysis
    sheet = workbook[sheet_names[0]]
    
    # Sample data from first few rows to look for indicators
    sample_data = []
    for r in range(1, min(20, sheet.max_row + 1)):
        row_data = []
        for c in range(1, min(15, sheet.max_column + 1)):
            cell_value = sheet.cell(row=r, column=c).value
            if cell_value:
                row_data.append(str(cell_value))
        sample_data.append(" ".join([str(x) for x in row_data if x]))
    
    sample_text = " ".join(sample_data)
    
    # Check for template indicators
    if "CEA" in sample_text and "Customer Engineering Approval" in sample_text:
        if sheet_names[0] == "CEA":
            return "CEA"
        
    if "Engineering Questionnaire" in sample_text or "EQ" in sample_text:
        if "STG Proposal" in sample_text and "Customer Decision" in sample_text:
            return "starteam"
        else:
            return "EQ"
    
    # If no clear match, try to determine by structure
    if sheet_names[0] == "CEA" and sheet.max_column <= 6:
        return "CEA"
    
    if sheet_names[0] == "EQ Template":
        if sheet.max_column >= 10:  # starteam has more columns
            return "starteam"
        else:
            return "EQ"
    
    # No match found
    return None

def extract_metadata(workbook, template_type):
    """
    Extract metadata fields based on template type
    Returns: Dictionary of metadata fields mapped to schema
    """
    metadata = {
        "customerName": "",
        "engineerName": None,
        "customerPN": [],
        "factoryPN": [],
        "stgPN": [],
        "baseMaterial": None,
        "solderMask": None,
        "viaPluggingType": None,
        "panelSize": None,
        "status": None,
        "stgSignatureDate": None,
        "stgSignatures": [],
        "customerSignatureDate": None,
        "customerSignatures": [],
        "createdAt": None  # Added to store the top date
    }
    
    sheet = workbook[workbook.sheetnames[0]]
    
    # Common extraction logic for all templates
    # Customer Name (row 1)
    if sheet.cell(row=1, column=3).value:
        metadata["customerName"] = str(sheet.cell(row=1, column=3).value)
    
    # Engineer Name (row 1)
    if sheet.cell(row=1, column=5).value:
        metadata["engineerName"] = str(sheet.cell(row=1, column=5).value)
    
    # Customer P/N (row 2)
    if sheet.cell(row=2, column=3).value:
        pn_value = str(sheet.cell(row=2, column=3).value)
        metadata["customerPN"] = [pn.strip() for pn in pn_value.split('\n') if pn.strip()]
    
    # Factory P/N (row 2)
    if sheet.cell(row=2, column=5).value:
        pn_value = str(sheet.cell(row=2, column=5).value)
        metadata["factoryPN"] = [pn.strip() for pn in pn_value.split('\n') if pn.strip()]
    
    # STG P/N (row 3)
    if sheet.cell(row=3, column=3).value:
        pn_value = str(sheet.cell(row=3, column=3).value)
        metadata["stgPN"] = [pn.strip() for pn in pn_value.split('\n') if pn.strip()]
    
    # Date (row 3, column 5) - This is the top date that should go to createdAt
    if sheet.cell(row=3, column=5).value:
        date_value = sheet.cell(row=3, column=5).value
        if hasattr(date_value, 'isoformat'):
            metadata["createdAt"] = date_value.isoformat()
    
    # Base Material and Solder Mask (row 7)
    if sheet.cell(row=7, column=3).value:
        metadata["baseMaterial"] = str(sheet.cell(row=7, column=3).value)
    
    if sheet.cell(row=7, column=5).value:
        metadata["solderMask"] = str(sheet.cell(row=7, column=5).value)
    
    # Template-specific extraction
    if template_type == "CEA":
        # Look for signature date and signatures at the bottom
        for row in range(15, min(25, sheet.max_row + 1)):  # Extended range to better catch signature info
            # Check for Date label
            if sheet.cell(row=row, column=1).value and "Date" in str(sheet.cell(row=row, column=1).value):
                # Check next row for actual date
                if sheet.cell(row=row+1, column=1).value:
                    try:
                        date_value = sheet.cell(row=row+1, column=1).value
                        if hasattr(date_value, 'isoformat'):
                            metadata["customerSignatureDate"] = date_value.isoformat()
                    except:
                        pass
            
            # Check for Customer's Signature label
            if sheet.cell(row=row, column=4).value and "Signature" in str(sheet.cell(row=row, column=4).value):
                # Check next row for actual signature
                if sheet.cell(row=row+1, column=4).value:
                    metadata["customerSignatures"] = [str(sheet.cell(row=row+1, column=4).value)]
    
    elif template_type == "EQ":
        # Look for signature info at the bottom
        for row in range(sheet.max_row - 10, sheet.max_row):
            # Check for signature related text
            if sheet.cell(row=row, column=1).value and ("Date" in str(sheet.cell(row=row, column=1).value) or "Signature" in str(sheet.cell(row=row, column=1).value)):
                # Check for date in next row
                if sheet.cell(row=row+1, column=1).value:
                    try:
                        date_value = sheet.cell(row=row+1, column=1).value
                        if hasattr(date_value, 'isoformat'):
                            metadata["customerSignatureDate"] = date_value.isoformat()
                    except:
                        pass
                
                # Check for signature in next row
                if sheet.cell(row=row+1, column=4).value:
                    metadata["customerSignatures"] = [str(sheet.cell(row=row+1, column=4).value)]
    
    elif template_type == "starteam":
        # Additional fields for starteam
        if sheet.cell(row=8, column=3).value:
            metadata["viaPluggingType"] = str(sheet.cell(row=8, column=3).value)
        
        if sheet.cell(row=8, column=5).value:
            metadata["panelSize"] = str(sheet.cell(row=8, column=5).value)
        
        # Look for signature info at the bottom
        for row in range(sheet.max_row - 10, sheet.max_row):
            cell_value = sheet.cell(row=row, column=4).value
            if cell_value and "Signature" in str(cell_value):
                # Check next row for actual signatures
                if sheet.cell(row=row+1, column=4).value:
                    stg_signature_text = str(sheet.cell(row=row+1, column=4).value)
                    # Split by newline to separate date and signature
                    lines = stg_signature_text.split('\n')
                    if len(lines) >= 2:
                        # First line is date, rest is signature
                        try:
                            from datetime import datetime
                            date_str = lines[0].strip()
                            # Try to parse the date
                            parsed_date = datetime.strptime(date_str, "%d.%m.%Y")
                            metadata["stgSignatureDate"] = parsed_date.isoformat()
                            metadata["stgSignatures"] = ['\n'.join(lines[1:]).strip()]
                        except:
                            # If date parsing fails, treat entire text as signature
                            metadata["stgSignatures"] = [stg_signature_text]
                    else:
                        metadata["stgSignatures"] = [stg_signature_text]
                
                if sheet.cell(row=row+1, column=5).value:
                    customer_signature_text = str(sheet.cell(row=row+1, column=5).value)
                    # Split by newline to separate date and signature
                    lines = customer_signature_text.split('\n')
                    if len(lines) >= 2:
                        # First line is date, rest is signature
                        try:
                            from datetime import datetime
                            date_str = lines[0].strip()
                            # Try to parse the date
                            parsed_date = datetime.strptime(date_str, "%d.%m.%Y")
                            metadata["customerSignatureDate"] = parsed_date.isoformat()
                            metadata["customerSignatures"] = ['\n'.join(lines[1:]).strip()]
                        except:
                            # If date parsing fails, treat entire text as signature
                            metadata["customerSignatures"] = [customer_signature_text]
                    else:
                        metadata["customerSignatures"] = [customer_signature_text]
    
    return metadata

def extract_questions(workbook, template_type, cell_to_images):
    """
    Extract questions based on template type
    Returns: List of question dictionaries
    """
    questions = []
    sheet = workbook[workbook.sheetnames[0]]
    
    # Get the created date from metadata for questions
    created_at = None
    if sheet.cell(row=3, column=5).value:
        date_value = sheet.cell(row=3, column=5).value
        if hasattr(date_value, 'isoformat'):
            created_at = date_value.isoformat()
    
    # Determine where questions start based on template type
    start_row = 9  # Default for CEA and EQ
    if template_type == "starteam":
        start_row = 10
    
    # Find the column indices for question data
    no_col, desc_col, sugg_col, resp_col = 1, 2, 3, 4  # Default for CEA
    
    if template_type == "EQ":
        no_col, desc_col, sugg_col, resp_col = 1, 2, 3, 4
    
    elif template_type == "starteam":
        no_col, desc_col, sugg_col, resp_col = 1, 2, 3, 5  # Customer Decision is in column 5
    
    # Extract questions
    current_row = start_row + 1  # Skip header row
    while current_row <= sheet.max_row:
        # Check if this is a question row (has a number in the first column)
        no_value = sheet.cell(row=current_row, column=no_col).value
        
        if no_value is not None and (isinstance(no_value, int) or (isinstance(no_value, str) and no_value.isdigit())):
            question = {
                "no": str(no_value),
                "description": None,
                "suggestion": None,
                "customerResponse": None,
                "descriptionImages": [],
                "suggestionImages": [],
                "customerResponseImages": [],
                "createdAt": created_at  # Use the top date for question createdAt
            }
            
            # Extract description with images
            desc_text, desc_images = get_cell_value_with_images(sheet, current_row, desc_col, cell_to_images)
            if desc_text and desc_text.strip():
                question["description"] = desc_text.strip()
            
            # Add all images found in the description cell
            for img_filename in desc_images:
                image_info = create_image_info({"filename": img_filename, "path": os.path.join("images", img_filename)})
                question["descriptionImages"].append(image_info)
                print(f"Added image {img_filename} to question {question['no']} description")
            
            # Extract suggestion with images
            sugg_text, sugg_images = get_cell_value_with_images(sheet, current_row, sugg_col, cell_to_images)
            if sugg_text and sugg_text.strip():
                question["suggestion"] = sugg_text.strip()
            
            # Add all images found in the suggestion cell
            for img_filename in sugg_images:
                image_info = create_image_info({"filename": img_filename, "path": os.path.join("images", img_filename)})
                question["suggestionImages"].append(image_info)
                print(f"Added image {img_filename} to question {question['no']} suggestion")
            
            # Extract customer response with images
            resp_text, resp_images = get_cell_value_with_images(sheet, current_row, resp_col, cell_to_images)
            if resp_text and resp_text.strip():
                question["customerResponse"] = resp_text.strip()
            
            # Add all images found in the customer response cell
            for img_filename in resp_images:
                image_info = create_image_info({"filename": img_filename, "path": os.path.join("images", img_filename)})
                question["customerResponseImages"].append(image_info)
                print(f"Added image {img_filename} to question {question['no']} customer response")
            
            questions.append(question)
        
        current_row += 1
    
    return questions

def extract_images(workbook, output_dir, template_type):
    """
    Extract images from the workbook and save to output directory
    Returns: Dictionary mapping cell coordinates to image filenames
    """
    images_dir = os.path.join(output_dir, "images")
    os.makedirs(images_dir, exist_ok=True)
    
    cell_to_images = {}
    image_counter = 1
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # Extract all images from the sheet
        for image in sheet._images:
            # Get the anchor information to determine which cell the image is in
            if hasattr(image, 'anchor') and hasattr(image.anchor, '_from'):
                # Get the cell coordinates where the image is anchored
                col = image.anchor._from.col
                row = image.anchor._from.row + 1  # openpyxl uses 0-based indexing, convert to 1-based
                
                # Convert column number to letter
                col_letter = chr(ord('A') + col)
                cell_coord = f"{col_letter}{row}"
                
                # Save the image
                image_filename = f"image_{image_counter}.png"
                image_path = os.path.join(images_dir, image_filename)
                
                try:
                    # Get image data
                    if hasattr(image, '_data'):
                        image_data = image._data()
                    else:
                        # Alternative method for getting image data
                        image_data = image.ref
                    
                    # Save image using PIL
                    pil_image = PILImage.open(io.BytesIO(image_data))
                    pil_image.save(image_path, 'PNG')
                    
                    # Store mapping with more precise positioning
                    if cell_coord not in cell_to_images:
                        cell_to_images[cell_coord] = []
                    cell_to_images[cell_coord].append(image_filename)
                    
                    image_counter += 1
                    print(f"Extracted image: {image_filename} from cell {cell_coord}")
                    
                except Exception as e:
                    print(f"Error extracting image {image_counter}: {e}")
                    image_counter += 1
    
    return cell_to_images

def get_cell_value_with_images(worksheet, row, col, cell_to_images):
    """
    Gets the cell value and associated images for a given row and column.
    Only returns images that are EXACTLY in this cell.
    """
    cell = worksheet.cell(row=row, column=col)
    cell_value = str(cell.value) if cell.value is not None else ""
    
    # Convert to cell coordinate
    col_letter = chr(ord('A') + col - 1)
    cell_coord = f"{col_letter}{row}"
    
    # Get images ONLY directly in this cell
    images = cell_to_images.get(cell_coord, []).copy()
    
    return cell_value, images

def associate_images_with_questions(questions, cell_to_images, workbook, template_type):
    """
    This function is now simplified since images are extracted directly in extract_questions.
    This can be used for any additional image association logic if needed.
    """
    # Images are now handled directly in extract_questions function
    # This function is kept for potential future enhancements
    return questions

def create_image_info(image_data):
    """Create image info object for schema"""
    image_id = generate_uuid()
    return {
        "id": image_id,
        "createdAt": datetime.now().isoformat(),
        "name": image_data["filename"],
        "type": "image/png",
        "key": generate_uuid(),
        "uploadUrl": image_data["path"]
    }

def map_to_schema(metadata, questions, file_name):
    """
    Map extracted data to the specified schema
    Returns: Dictionary conforming to schema
    """
    # Create File object for original file
    original_file = {
        "id": generate_uuid(),
        "createdAt": datetime.now().isoformat(),
        "name": file_name,
        "type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "key": generate_uuid(),
        "uploadUrl": file_name
    }
    
    # Create EQ object
    eq = {
        "id": generate_uuid(),
        "createdAt": metadata.get("createdAt") or datetime.now().isoformat(),  # Use top date if available
        "customerName": metadata.get("customerName", ""),
        "engineerName": metadata.get("engineerName"),
        "customerPN": metadata.get("customerPN", []),
        "factoryPN": metadata.get("factoryPN", []),
        "stgPN": metadata.get("stgPN", []),
        "baseMaterial": metadata.get("baseMaterial"),
        "solderMask": metadata.get("solderMask"),
        "viaPluggingType": metadata.get("viaPluggingType"),
        "panelSize": metadata.get("panelSize"),
        "status": "Closed",  # Default status
        "stgSignatureDate": metadata.get("stgSignatureDate"),
        "stgSignatures": metadata.get("stgSignatures", []),
        "customerSignatureDate": metadata.get("customerSignatureDate"),
        "customerSignatures": metadata.get("customerSignatures", []),
        "fileName": file_name,
        "originalFile": original_file,
        "questions": []
    }
    
    # Map questions to schema
    for q in questions:
        question = {
            "id": generate_uuid(),
            "createdAt": q.get("createdAt") or datetime.now().isoformat(),  # Use top date if available
            "no": q.get("no", ""),
            "description": q.get("description"),
            "suggestion": q.get("suggestion"),
            "customerResponse": q.get("customerResponse"),
            "descriptionImages": q.get("descriptionImages", []),
            "suggestionImages": q.get("suggestionImages", []),
            "customerResponseImages": q.get("customerResponseImages", [])
        }
        eq["questions"].append(question)
    
    return eq

def validate_extracted_data(eq_data):
    """
    Validate extracted data against schema requirements
    Returns: Boolean indicating if valid, and list of validation errors
    """
    errors = []
    
    # Check required fields
    if not eq_data.get("customerName"):
        errors.append("Missing required field: customerName")
    
    if not eq_data.get("fileName"):
        errors.append("Missing required field: fileName")
    
    # Validate questions
    for i, question in enumerate(eq_data.get("questions", [])):
        if not question.get("no"):
            errors.append(f"Question {i+1} missing required field: no")
    
    return len(errors) == 0, errors

def handle_outlier_file(file_path, outlier_dir):
    """
    Handle files that don't match any of the 3 templates by copying them to outlier directory
    Returns: Boolean indicating success
    """
    try:
        file_name = os.path.basename(file_path)
        outlier_file_path = os.path.join(outlier_dir, file_name)
        
        # Copy the file to outlier directory
        shutil.copy2(file_path, outlier_file_path)
        
        # Create a simple info file about the outlier
        info_file_path = os.path.join(outlier_dir, f"{os.path.splitext(file_name)[0]}_info.txt")
        with open(info_file_path, "w", encoding='utf-8') as f:
            f.write(f"Outlier File: {file_name}\n")
            f.write(f"Original Path: {file_path}\n")
            f.write(f"Processed Date: {datetime.now().isoformat()}\n")
            f.write("Reason: Does not match any of the 3 supported templates (CEA, EQ, starteam)\n")
        
        print(f"Moved outlier file {file_name} to outlier directory")
        return True
    
    except Exception as e:
        print(f"Error handling outlier file {file_path}: {str(e)}")
        return False

def process_excel_file(file_path, output_dir, outlier_dir):
    """
    Process a single Excel file and generate output
    Returns: Tuple (success: Boolean, is_outlier: Boolean)
    """
    try:
        # Get file name
        file_name = os.path.basename(file_path)
        
        # Load workbook
        workbook = load_workbook(file_path, data_only=True)
        
        # Identify template type
        template_type = identify_template_type(workbook)
        if not template_type:
            print(f"File {file_path} does not match any template - moving to outlier directory")
            success = handle_outlier_file(file_path, outlier_dir)
            return success, True  # Return (success, is_outlier)
        
        print(f"Processing {file_path} as {template_type} template")
        
        # Create output directory
        file_output_dir = os.path.join(output_dir, os.path.basename(file_path).split('.')[0])
        os.makedirs(os.path.join(file_output_dir, "images"), exist_ok=True)
        
        # Extract data
        metadata = extract_metadata(workbook, template_type)
        cell_to_images = extract_images(workbook, file_output_dir, template_type)
        questions = extract_questions(workbook, template_type, cell_to_images)
        
        # Images are now precisely associated during extraction
        # questions = associate_images_with_questions(questions, cell_to_images, workbook, template_type)
        
        # Print summary of extracted images
        total_images = sum(len(images) for images in cell_to_images.values())
        print(f"Total images extracted: {total_images}")
        for cell_coord, images in cell_to_images.items():
            if images:
                print(f"Cell {cell_coord}: {len(images)} image(s) - {', '.join(images)}")
        
        # Map to schema
        eq_data = map_to_schema(metadata, questions, file_name)
        
        # Validate
        is_valid, errors = validate_extracted_data(eq_data)
        if not is_valid:
            print(f"Validation errors in {file_path}: {errors}")
            return False, False  # Return (success, is_outlier)
        
        # Write output
        with open(os.path.join(file_output_dir, "index.json"), "w", encoding='utf-8') as f:
            json.dump(eq_data, f, indent=2, ensure_ascii=False)
        
        # Copy the original Excel file with UUID as filename
        excel_copy_path = os.path.join(file_output_dir, f"index.xlsx")
        try:
            shutil.copy2(file_path, excel_copy_path)
            print(f"Copied original Excel file to: {excel_copy_path}")
        except Exception as e:
            print(f"Warning: Failed to copy Excel file: {str(e)}")
        
        print(f"Successfully processed {file_path}")
        return True, False  # Return (success, is_outlier)
    
    except Exception as e:
        print(f"Error processing {file_path}: {str(e)}")
        import traceback
        traceback.print_exc()
        return False, False  # Return (success, is_outlier)

def main():
    """
    Main entry point for the script
    """
    # Set up argument parser
    parser = argparse.ArgumentParser(description='Convert Excel files to JSON format based on template types')
    parser.add_argument('--input-dir', '-i', default='./raw_data', 
                       help='Input directory containing Excel files (default: ./raw_data)')
    parser.add_argument('--output-dir', '-o', default='./processed_data',
                       help='Output directory for processed JSON files (default: ./processed_data)')
    parser.add_argument('--outlier-dir', '-l', default='./outlier_files',
                       help='Directory for files that don\'t match any template (default: ./outlier_files)')
    
    args = parser.parse_args()
    
    # Use arguments instead of hardcoded paths
    raw_data_dir = args.input_dir
    output_dir = args.output_dir
    outlier_dir = args.outlier_dir
    
    # Create output directories
    os.makedirs(output_dir, exist_ok=True)
    os.makedirs(outlier_dir, exist_ok=True)
    
    # Check if input directory exists
    if not os.path.exists(raw_data_dir):
        print(f"Error: Input directory '{raw_data_dir}' does not exist")
        return
    
    # Get all xlsx files from input directory
    excel_files = [os.path.join(raw_data_dir, f) for f in os.listdir(raw_data_dir) 
                  if f.endswith(".xlsx") or f.endswith(".xls")]
    
    if not excel_files:
        print(f"No Excel files found in {raw_data_dir}")
        return
    
    print(f"Input directory: {raw_data_dir}")
    print(f"Output directory: {output_dir}")
    print(f"Outlier directory: {outlier_dir}")
    print(f"Found {len(excel_files)} Excel files to process\n")
    
    # Process each file
    success_count = 0
    outlier_count = 0
    processed_count = 0
    
    for file_path in excel_files:
        success, is_outlier = process_excel_file(file_path, output_dir, outlier_dir)
        if success:
            success_count += 1
            if is_outlier:
                outlier_count += 1
            else:
                processed_count += 1
    
    print(f"\n=== Processing Summary ===")
    print(f"Total files: {len(excel_files)}")
    print(f"Successfully processed (matching templates): {processed_count}")
    print(f"Outlier files (moved to outlier directory): {outlier_count}")
    print(f"Failed to process: {len(excel_files) - success_count}")
    print(f"Overall success rate: {success_count}/{len(excel_files)} files")

if __name__ == "__main__":
    main()
